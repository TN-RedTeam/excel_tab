"""Import d'un classeur existant (§9.4) et persistance SQLite."""

from __future__ import annotations

import datetime as dt
from decimal import Decimal

import openpyxl
import pytest

from tpt_app.core import moteur
from tpt_app.core.arrondi import arrondi_centime
from tpt_app.core.models import Dossier, REGIME_ML36, REGIME_ML37
from tpt_app.db.repository import DepotDossiers, deserialiser, serialiser
from tpt_app.export import excel
from tpt_app.export.excel import CHEMIN_TEMPLATE
from tpt_app.importer.classeur import ImportImpossible, importer
from tpt_app.mapping_classeur import FEUILLE_ML36


# --------------------------------------------------------------------------
# Import
# --------------------------------------------------------------------------


def test_aller_retour_export_puis_import(dossier_test2, tmp_path):
    """Un dossier exporté puis réimporté produit exactement les mêmes montants."""
    resultat = moteur.calculer(dossier_test2)
    chemin = excel.exporter(dossier_test2, resultat, tmp_path / "dossier.xlsx")

    reimporte = importer(chemin)
    assert reimporte.regime == REGIME_ML37
    assert reimporte.ml37.salarie.nom == "DUPONT"
    assert reimporte.ml37.tmf_100 == Decimal(3000)
    assert reimporte.ml37.nb_jours_mois == 31
    assert reimporte.ml37.taux_tpt == Decimal("0.5")

    apres = moteur.calculer(reimporte)
    for avant_periode, apres_periode in zip(resultat.ml37.periodes, apres.ml37.periodes):
        assert arrondi_centime(avant_periode.montant_declare) == \
            arrondi_centime(apres_periode.montant_declare)
    assert arrondi_centime(apres.ml37.perte_cpam) == \
        arrondi_centime(resultat.ml37.perte_cpam)


def test_import_lit_les_deux_lignes_de_saisie(tmp_path):
    """Les dates saisies sur la ligne « motif » sont reprises sans perte (§5.3)."""
    classeur = openpyxl.load_workbook(CHEMIN_TEMPLATE)
    feuille = classeur[FEUILLE_ML36]
    feuille["B4"] = "MARTIN"
    feuille["B3"] = "B98765"
    feuille["F2"] = 30
    feuille["B10"] = 2500

    # Période 1 : dates sur la ligne « période » (lignes 21/22).
    feuille["A21"] = "ML36"
    feuille["B21"] = dt.datetime(2025, 7, 1)
    feuille["C21"] = dt.datetime(2025, 7, 8)
    # Période 2 : motif d'absence, dates portées par la ligne « motif ».
    feuille["A27"] = "Maladie"
    feuille["B27"] = dt.datetime(2025, 7, 11)
    feuille["C27"] = dt.datetime(2025, 7, 18)

    chemin = tmp_path / "historique.xlsx"
    classeur.save(chemin)

    dossier = importer(chemin)
    premiere, seconde = dossier.ml36.periodes[0], dossier.ml36.periodes[1]

    assert premiere.date_debut == dt.date(2025, 7, 1)
    assert premiere.dates_sur_ligne_periode is True

    assert seconde.motif_absence == "Maladie"
    assert seconde.date_debut == dt.date(2025, 7, 11)
    assert seconde.date_fin == dt.date(2025, 7, 18)
    # La période d'absence ne compte pas dans les 30èmes.
    assert seconde.sur_ligne_periode is False
    assert seconde.nb_jours_ligne_periode() == 0


def test_import_conserve_le_calcul_d_origine(tmp_path):
    """Dates d'absence sur la ligne « période » : le mode v6 les valorise."""
    classeur = openpyxl.load_workbook(CHEMIN_TEMPLATE)
    feuille = classeur[FEUILLE_ML36]
    feuille["B4"] = "MARTIN"
    feuille["F2"] = 30
    feuille["B10"] = 2500
    feuille["D8"] = 0.4
    feuille["B8"] = 1
    feuille["A21"] = "ML36"
    feuille["A22"] = "Maladie"
    feuille["B21"] = dt.datetime(2025, 7, 1)      # dates sur la « mauvaise » ligne
    feuille["C21"] = dt.datetime(2025, 7, 5)
    chemin = tmp_path / "ambigu.xlsx"
    classeur.save(chemin)

    dossier = importer(chemin)
    periode = dossier.ml36.periodes[0]
    assert periode.dates_sur_ligne_periode is True
    assert periode.est_absence

    resultat = moteur.calculer(dossier)
    assert arrondi_centime(resultat.ml36.periodes[0].montant_declare) == Decimal("166.67")
    # L'attestation masque néanmoins le montant derrière le motif.
    assert resultat.attestation.lignes[0].libelle == "Maladie"


def test_import_refuse_un_fichier_etranger(tmp_path):
    classeur = openpyxl.Workbook()
    classeur.active["A1"] = "rien à voir"
    chemin = tmp_path / "autre.xlsx"
    classeur.save(chemin)

    with pytest.raises(ImportImpossible, match="onglets attendus"):
        importer(chemin)


def test_import_refuse_un_fichier_illisible(tmp_path):
    chemin = tmp_path / "corrompu.xlsx"
    chemin.write_bytes(b"ceci n'est pas un classeur")
    with pytest.raises(ImportImpossible):
        importer(chemin)


# --------------------------------------------------------------------------
# Persistance
# --------------------------------------------------------------------------


def test_serialisation_conserve_les_decimaux_et_les_dates(dossier_test1):
    reconstruit = deserialiser(serialiser(dossier_test1))

    assert reconstruit.regime == dossier_test1.regime
    assert reconstruit.ml36.tmf_100 == Decimal(2500)
    assert isinstance(reconstruit.ml36.taux_tpt, Decimal)
    assert reconstruit.ml36.mois == dt.date(2025, 7, 1)
    assert reconstruit.ml36.periodes[0].date_debut == dt.date(2025, 7, 1)
    assert reconstruit.ml36.periodes[1].motif_absence == "Maladie"

    # Les résultats recalculés sont rigoureusement identiques.
    avant = moteur.calculer(dossier_test1)
    apres = moteur.calculer(reconstruit)
    assert arrondi_centime(apres.ml36.perte_cpam) == arrondi_centime(avant.ml36.perte_cpam)


def test_depot_enregistre_charge_et_supprime(dossier_test1, tmp_path):
    with DepotDossiers(tmp_path / "app.db") as depot:
        dossier_test1.libelle = "DUPONT — juillet 2025"
        identifiant = depot.enregistrer(dossier_test1)
        assert identifiant == dossier_test1.identifiant

        recharge = depot.charger(identifiant)
        assert recharge is not None
        assert recharge.libelle == "DUPONT — juillet 2025"
        assert recharge.ml36.tmf_100 == Decimal(2500)

        # Mise à jour en place, sans créer de doublon.
        recharge.ml36.tmf_100 = Decimal(2600)
        depot.enregistrer(recharge)
        assert len(depot.lister()) == 1
        assert depot.charger(identifiant).ml36.tmf_100 == Decimal(2600)

        depot.supprimer(identifiant)
        assert depot.lister() == []
        assert depot.charger(identifiant) is None


def test_depot_filtre_la_liste(dossier_test1, dossier_test2, tmp_path):
    with DepotDossiers(tmp_path / "app.db") as depot:
        depot.enregistrer(dossier_test1)
        dossier_test2.ml37.salarie.nom = "BERNARD"
        dossier_test2.ml37.salarie.matricule = "C55555"
        depot.enregistrer(dossier_test2)

        assert len(depot.lister()) == 2
        assert [d["nom"] for d in depot.lister(recherche="BERNARD")] == ["BERNARD"]
        assert [d["nom"] for d in depot.lister(recherche="A12345")] == ["DUPONT"]
        assert [d["regime"] for d in depot.lister(regime=REGIME_ML36)] == [REGIME_ML36]
        assert len(depot.lister(mois="2025-07")) == 2
        assert depot.lister(mois="2025-08") == []
