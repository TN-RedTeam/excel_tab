"""§8 — Test 6 : fidélité des exports."""

from __future__ import annotations

import openpyxl
import pytest

from tpt_app.core import moteur
from tpt_app.core.attestation import nom_fichier
from tpt_app.core.models import REGIME_ML36, Periode
from tpt_app.export import excel, pdf
from tpt_app.export.excel import CHEMIN_TEMPLATE
from tpt_app.mapping_classeur import FEUILLE_ATTESTATION

ATTRIBUTS_STYLE = ("font", "fill", "border", "alignment", "number_format", "protection")


def normaliser(objet):
    """Réduit un objet de style openpyxl à des types comparables.

    Les classes de style d'openpyxl n'implémentent pas ``__eq__`` : deux
    chargements du même fichier produisent des objets distincts. On compare donc
    leur structure, récursivement.

    ``False`` est assimilé à ``None`` : les booléens optionnels d'OOXML
    (``shrinkToFit``, ``bold``…) sont réécrits par openpyxl sous leur forme
    absente lors d'un simple aller-retour, sans écriture de valeur, et les deux
    formes ont exactement le même effet dans Excel.
    """
    if objet is False:
        return None
    if objet is None or isinstance(objet, (str, int, float, bool)):
        return objet
    if isinstance(objet, (list, tuple)):
        return [normaliser(element) for element in objet]
    attributs = getattr(objet, "__dict__", None)
    if attributs:
        return {clef: normaliser(valeur) for clef, valeur in sorted(attributs.items())}
    return str(objet)


@pytest.fixture
def classeur_exporte(dossier_test2, tmp_path):
    resultat = moteur.calculer(dossier_test2)
    chemin = excel.exporter(dossier_test2, resultat, tmp_path / "attestation.xlsx")
    return openpyxl.load_workbook(chemin)


def test6_styles_identiques_au_template(classeur_exporte):
    """Aucune différence de style, de fusion ou de dimension avec le template."""
    template = openpyxl.load_workbook(CHEMIN_TEMPLATE)

    assert classeur_exporte.sheetnames == template.sheetnames

    for nom in template.sheetnames:
        attendu, obtenu = template[nom], classeur_exporte[nom]

        assert {str(m) for m in obtenu.merged_cells.ranges} == \
               {str(m) for m in attendu.merged_cells.ranges}, nom
        assert obtenu.print_area == attendu.print_area, nom
        assert obtenu.page_setup.orientation == attendu.page_setup.orientation, nom
        assert obtenu.page_setup.paperSize == attendu.page_setup.paperSize, nom

        assert {c: d.width for c, d in obtenu.column_dimensions.items()} == \
               {c: d.width for c, d in attendu.column_dimensions.items()}, nom
        assert {r: d.height for r, d in obtenu.row_dimensions.items()} == \
               {r: d.height for r, d in attendu.row_dimensions.items()}, nom

        for ligne in attendu.iter_rows():
            for cellule in ligne:
                # Les cellules absentes du XML du template n'ont pas de style à
                # préserver : Excel leur applique de toute façon le format par
                # défaut du classeur. On vérifie que toutes les autres — celles
                # qui portent réellement une mise en forme — sont intactes.
                if cellule._style is None:
                    continue
                autre = obtenu[cellule.coordinate]
                for attribut in ATTRIBUTS_STYLE:
                    assert normaliser(getattr(autre, attribut)) == \
                        normaliser(getattr(cellule, attribut)), \
                        f"{nom}!{cellule.coordinate}.{attribut}"


def test6_images_conservees(classeur_exporte):
    template = openpyxl.load_workbook(CHEMIN_TEMPLATE)
    for nom in template.sheetnames:
        assert len(classeur_exporte[nom]._images) == len(template[nom]._images), nom


def test6_bandeau_orange(classeur_exporte):
    """Le bandeau « IDENTIFICATION DU SALARIE » reste #FF6600 sur texte blanc."""
    cellule = classeur_exporte[FEUILLE_ATTESTATION]["B9"]
    assert cellule.value == "IDENTIFICATION DU SALARIE"
    assert cellule.fill.fgColor.rgb.endswith("FF6600")
    assert cellule.font.color.rgb.endswith("FFFFFF")
    assert cellule.font.bold


def test6_valeurs_ecrites_sans_formule(classeur_exporte):
    """Les formules sont remplacées par des valeurs : le fichier est auditable."""
    feuille = classeur_exporte[FEUILLE_ATTESTATION]
    assert feuille["D11"].value == "DUPONT"
    assert feuille["D26"].value is not None
    for ligne in classeur_exporte[FEUILLE_ATTESTATION].iter_rows():
        for cellule in ligne:
            assert not (isinstance(cellule.value, str) and cellule.value.startswith("=")), \
                cellule.coordinate


def test6_pdf_une_seule_page_a4(dossier_test2, tmp_path):
    pypdfium2 = pytest.importorskip("pypdfium2")
    resultat = moteur.calculer(dossier_test2)
    chemin = pdf.exporter(dossier_test2, resultat, tmp_path / "attestation.pdf")

    document = pypdfium2.PdfDocument(chemin)
    assert len(document) == 1
    largeur, hauteur = document[0].get_size()
    assert round(largeur) == 595 and round(hauteur) == 842   # A4 portrait


def test6_pdf_contient_les_montants(dossier_test2, tmp_path):
    pypdfium2 = pytest.importorskip("pypdfium2")
    resultat = moteur.calculer(dossier_test2)
    chemin = pdf.exporter(dossier_test2, resultat, tmp_path / "attestation.pdf")

    texte = pypdfium2.PdfDocument(chemin)[0].get_textpage().get_text_range()
    for attendu in ("ATTESTATION DE TEMPS PARTIEL", "DUPONT", "01/07/2025",
                    "Congés annuels", "Maladie", "Absence sans solde",
                    "78,13", "171,87", "56,25", "123,75", "50,00 %"):
        assert attendu in texte, attendu


def test_nommage_des_fichiers(dossier_test2):
    resultat = moteur.calculer(dossier_test2)
    assert nom_fichier(resultat.attestation, dossier_test2.ml37.mois, "pdf") == \
        "ATTESTATION_VIVINTER_DUPONT_A12345_2025-07.pdf"
    assert nom_fichier(resultat.attestation, dossier_test2.ml37.mois, "xlsx") == \
        "ATTESTATION_VIVINTER_DUPONT_A12345_2025-07.xlsx"


def bordures_cloture_attendues():
    """Tramage de la ligne de clôture du tableau, relevé dans le gabarit."""
    template = openpyxl.load_workbook(CHEMIN_TEMPLATE)[FEUILLE_ATTESTATION]
    return [
        tuple(bool(getattr(template.cell(row=32, column=colonne).border, cote).style)
              for cote in ("left", "right", "top", "bottom"))
        for colonne in range(2, 9)
    ]


def test_export_declare_toutes_les_periodes(dossier_test1, tmp_path):
    """Au-delà de 7 périodes, le tableau s'étend au lieu de bloquer l'export."""
    dossier_test1.ml36.nb_jours_mois = 31
    dossier_test1.ml36.periodes = [
        Periode(motif_principal=REGIME_ML36,
                date_debut=dossier_test1.ml36.mois.replace(day=1 + 3 * i),
                date_fin=dossier_test1.ml36.mois.replace(day=3 + 3 * i))
        for i in range(10)
    ]
    resultat = moteur.calculer(dossier_test1)
    assert resultat.exportable

    chemin = excel.exporter(dossier_test1, resultat, tmp_path / "dix.xlsx")
    feuille = openpyxl.load_workbook(chemin)[FEUILLE_ATTESTATION]

    def bordures(ligne):
        return [
            tuple(bool(getattr(feuille.cell(row=ligne, column=colonne).border, cote).style)
                  for cote in ("left", "right", "top", "bottom"))
            for colonne in range(2, 9)
        ]

    # 10 lignes de période renseignées, à partir de la ligne 26.
    for rang in range(10):
        assert feuille.cell(row=26 + rang, column=2).value is not None, rang

    # Les lignes ajoutées (32 à 34) reprennent exactement le tramage de la ligne
    # courante du gabarit ; la ligne 35 reste la ligne de clôture du tableau.
    modele = bordures(31)
    for ligne in (32, 33, 34):
        assert bordures(ligne) == modele, f"ligne {ligne}"
    assert bordures(35) == bordures_cloture_attendues()
    # Rien n'est dessiné sous le tableau.
    assert bordures(36) == [(False,) * 4] * 7

    # Les champs du bas ont suivi le décalage de 3 lignes.
    assert feuille["C39"].value == "ROISSY CDG"
    assert feuille["F39"].value == "Cachet et Signature"


def test_export_pdf_tient_sur_une_page_avec_dix_periodes(dossier_test1, tmp_path):
    pypdfium2 = pytest.importorskip("pypdfium2")

    dossier_test1.ml36.nb_jours_mois = 31
    dossier_test1.ml36.periodes = [
        Periode(motif_principal=REGIME_ML36,
                date_debut=dossier_test1.ml36.mois.replace(day=1 + 3 * i),
                date_fin=dossier_test1.ml36.mois.replace(day=3 + 3 * i))
        for i in range(10)
    ]
    resultat = moteur.calculer(dossier_test1)
    chemin = pdf.exporter(dossier_test1, resultat, tmp_path / "dix.pdf")

    document = pypdfium2.PdfDocument(chemin)
    assert len(document) == 1
    largeur, hauteur = document[0].get_size()
    assert round(largeur) == 595 and round(hauteur) == 842

    texte = document[0].get_textpage().get_text_range()
    # Les dix périodes figurent bien, ainsi que le bas de page.
    for jour in range(1, 30, 3):
        assert f"{jour:02d}/07/2025" in texte
    assert "Cachet et Signature" in texte
    assert "A retourner à VIVINTER" in texte
