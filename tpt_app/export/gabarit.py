"""Extension du tableau des périodes de l'attestation.

Le gabarit Vivinter comporte 7 lignes de période (26 à 32). Lorsqu'un dossier en
compte davantage, ce module insère les lignes manquantes dans la feuille chargée,
en reproduisant à l'identique la mise en forme d'une ligne courante.

``Worksheet.insert_rows`` décale les cellules et leurs styles, mais **ni les
fusions, ni les hauteurs de ligne, ni la zone d'impression** : ce module compense
chacun de ces points. Les exports Excel et PDF consomment la même feuille étendue,
ce qui garantit deux documents identiques.
"""

from __future__ import annotations

from copy import copy

from openpyxl.worksheet.cell_range import CellRange

#: Première ligne de période du gabarit.
LIGNE_PREMIERE_PERIODE = 26
#: Nombre de lignes de période du gabarit d'origine.
NB_LIGNES_MODELE = 7
#: Dernière ligne de période : elle porte la bordure inférieure du tableau.
LIGNE_CLOTURE = LIGNE_PREMIERE_PERIODE + NB_LIGNES_MODELE - 1      # 32
#: Ligne courante servant de modèle aux lignes ajoutées (bordures complètes).
LIGNE_MODELE = LIGNE_CLOTURE - 1                                    # 31

#: Dernière ligne de la zone d'impression du gabarit (``A1:I55``).
DERNIERE_LIGNE_IMPRESSION = 55
COLONNE_IMPRESSION = "I"


def _decaler_hauteurs(feuille, insertion: int, supplement: int) -> None:
    """Redescend les hauteurs de ligne et attribue celle du modèle aux nouvelles."""
    hauteurs = {
        ligne: dimension.height
        for ligne, dimension in feuille.row_dimensions.items()
        if dimension.height
    }
    hauteur_modele = hauteurs.get(LIGNE_MODELE)

    nouvelles = {
        (ligne + supplement if ligne >= insertion else ligne): hauteur
        for ligne, hauteur in hauteurs.items()
    }
    for offset in range(supplement):
        nouvelles[insertion + offset] = hauteur_modele

    for ligne, hauteur in nouvelles.items():
        feuille.row_dimensions[ligne].height = hauteur


def _decaler_fusions(feuille, insertion: int, supplement: int) -> None:
    """Descend les fusions situées sous le point d'insertion.

    Les plages sont décalées **en place**. Passer par ``unmerge_cells`` puis
    ``merge_cells`` réinitialiserait le style de toutes les cellules démergées :
    comme les fusions n'ont pas encore bougé au moment du démergeage, elles
    recouvrent les lignes fraîchement insérées et effaceraient leur mise en
    forme.
    """
    for plage in feuille.merged_cells.ranges:
        if plage.min_row >= insertion:
            plage.shift(row_shift=supplement)


def _fusions_du_modele(feuille) -> list[CellRange]:
    return [
        CellRange(str(m)) for m in feuille.merged_cells.ranges
        if m.min_row == LIGNE_MODELE and m.max_row == LIGNE_MODELE
    ]


def uniformiser_lignes_periode(feuille, nb_lignes: int) -> None:
    """Aligne toutes les lignes de période sur la mise en forme de la ligne modèle.

    Le gabarit fourni comporte deux irrégularités sur sa 2ème ligne de période :
    ``G27`` a conservé la police par défaut au lieu de l'Arial 10 gras du reste de
    la colonne « Autres primes », et ``B27``/``C27`` ont perdu leurs bordures
    horizontales. Elles se voient sur toute attestation d'au moins deux périodes.

    Les lignes 26 à l'avant-dernière reçoivent donc le style de la ligne modèle ;
    la ligne de clôture garde le sien, elle seule portant la bordure inférieure.
    """
    styles = {
        colonne: feuille.cell(row=LIGNE_MODELE, column=colonne)._style
        for colonne in range(2, 9)          # colonnes B à H
    }
    derniere = LIGNE_PREMIERE_PERIODE + max(nb_lignes, NB_LIGNES_MODELE) - 1
    for ligne in range(LIGNE_PREMIERE_PERIODE, derniere):
        for colonne, style in styles.items():
            feuille.cell(row=ligne, column=colonne)._style = copy(style)


def etendre_tableau_periodes(feuille, nb_lignes: int) -> int:
    """Porte le tableau des périodes à ``nb_lignes`` lignes.

    Renvoie le décalage appliqué à tout ce qui se trouve sous le tableau, à
    reporter sur les coordonnées des champs du bas (``Fait à``, ``le``, ``Mail``…).
    """
    supplement = max(0, nb_lignes - NB_LIGNES_MODELE)
    if supplement == 0:
        return 0

    modeles = _fusions_du_modele(feuille)
    styles_modele = {
        colonne: copy(feuille.cell(row=LIGNE_MODELE, column=colonne)._style)
        for colonne in range(1, feuille.max_column + 1)
    }

    feuille.insert_rows(LIGNE_CLOTURE, supplement)

    # L'ordre compte : les fusions sont d'abord replacées, puis les lignes
    # ajoutées sont peintes, et enfin fusionnées à leur tour.
    _decaler_fusions(feuille, LIGNE_CLOTURE, supplement)
    _decaler_hauteurs(feuille, LIGNE_CLOTURE, supplement)

    for offset in range(supplement):
        cible = LIGNE_CLOTURE + offset
        for colonne, style in styles_modele.items():
            feuille.cell(row=cible, column=colonne)._style = copy(style)

    # Les fusions internes à une ligne de période (E:F) sont reproduites.
    for offset in range(supplement):
        cible = LIGNE_CLOTURE + offset
        for plage in modeles:
            feuille.merge_cells(
                start_row=cible, start_column=plage.min_col,
                end_row=cible, end_column=plage.max_col,
            )

    derniere = DERNIERE_LIGNE_IMPRESSION + supplement
    feuille.print_area = f"A1:{COLONNE_IMPRESSION}{derniere}"
    return supplement


def plage_impression(supplement: int) -> str:
    """Zone d'impression correspondant au nombre de lignes ajoutées."""
    return f"A1:{COLONNE_IMPRESSION}{DERNIERE_LIGNE_IMPRESSION + supplement}"


def decaler(coordonnee: str, supplement: int) -> str:
    """Décale une coordonnée si elle se situe sous le tableau des périodes."""
    if supplement == 0:
        return coordonnee
    lettres = "".join(c for c in coordonnee if c.isalpha())
    ligne = int("".join(c for c in coordonnee if c.isdigit()))
    if ligne >= LIGNE_CLOTURE:
        ligne += supplement
    return f"{lettres}{ligne}"


def ligne_periode(index: int) -> int:
    """Ligne de la période d'indice ``index`` (0-based) dans la feuille étendue."""
    return LIGNE_PREMIERE_PERIODE + index


#: Onglet « mode d'emploi » du classeur d'origine. Il n'a aucun rôle dans une
#: attestation et pèse à lui seul près de 90 % du poids du fichier produit
#: (4 images, ~280 Ko) : il est retiré des exports.
FEUILLE_MODE_EMPLOI = "MODOP Attestation VIV"


def alleger_classeur(classeur) -> None:
    """Retire du classeur exporté ce qui ne sert pas à l'attestation."""
    if FEUILLE_MODE_EMPLOI in classeur.sheetnames:
        del classeur[FEUILLE_MODE_EMPLOI]
