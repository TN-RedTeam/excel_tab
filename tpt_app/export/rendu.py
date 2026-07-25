"""Rendu ReportLab d'une plage de feuille Excel.

Plutôt que de retranscrire à la main la charte de l'attestation, ce module lit la
géométrie et les styles directement dans le template embarqué (largeurs de
colonnes, hauteurs de lignes, fusions, bordures, polices, couleurs, alignements)
et les restitue au trait près sur un canevas ReportLab. Le texte des cellules
calculées est fourni par l'appelant, déjà formaté en français.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional

from openpyxl.utils import get_column_letter, range_boundaries
from reportlab.lib.colors import Color, HexColor, black
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as rl_canvas

DOSSIER_POLICES = Path(__file__).resolve().parent / "assets" / "fonts"

#: Largeur Excel par défaut, en « caractères ».
LARGEUR_COLONNE_DEFAUT = 8.43
#: Hauteur de ligne Excel par défaut, en points.
HAUTEUR_LIGNE_DEFAUT = 15.0

_POLICES_ENREGISTREES = False


def _chemins_candidats(nom_fichier: str) -> Iterable[Path]:
    yield DOSSIER_POLICES / nom_fichier
    for racine in ("C:/Windows/Fonts", "/usr/share/fonts/truetype/liberation",
                   "/usr/share/fonts/truetype/msttcorefonts",
                   "/Library/Fonts", "/System/Library/Fonts/Supplemental"):
        yield Path(racine) / nom_fichier


def _enregistrer(nom_reportlab: str, *fichiers: str) -> bool:
    for fichier in fichiers:
        for chemin in _chemins_candidats(fichier):
            if chemin.is_file():
                try:
                    pdfmetrics.registerFont(TTFont(nom_reportlab, str(chemin)))
                    return True
                except Exception:      # police illisible : on tente la suivante
                    continue
    return False


def enregistrer_polices() -> dict[tuple[str, bool], str]:
    """Enregistre les polices et renvoie la table de correspondance Excel → PDF.

    Arial est utilisée si elle est présente (poste Windows) ; à défaut Liberation
    Sans, métriquement identique à Arial et embarquée dans le bundle ; en dernier
    recours Helvetica, également compatible avec les métriques Arial.
    """
    global _POLICES_ENREGISTREES
    if not _POLICES_ENREGISTREES:
        _enregistrer("ArialTPT", "arial.ttf", "Arial.ttf", "LiberationSans-Regular.ttf")
        _enregistrer("ArialTPT-Bold", "arialbd.ttf", "Arial Bold.ttf",
                     "LiberationSans-Bold.ttf")
        _enregistrer("ArialNarrowTPT", "arialn.ttf", "LiberationSansNarrow-Regular.ttf",
                     "arial.ttf", "LiberationSans-Regular.ttf")
        _enregistrer("ArialNarrowTPT-Bold", "arialnb.ttf",
                     "LiberationSansNarrow-Bold.ttf", "arialbd.ttf",
                     "LiberationSans-Bold.ttf")
        _POLICES_ENREGISTREES = True

    connues = set(pdfmetrics.getRegisteredFontNames())

    def choisir(prefere: str, secours: str) -> str:
        return prefere if prefere in connues else secours

    normale = choisir("ArialTPT", "Helvetica")
    grasse = choisir("ArialTPT-Bold", "Helvetica-Bold")
    etroite = choisir("ArialNarrowTPT", normale)
    etroite_grasse = choisir("ArialNarrowTPT-Bold", grasse)

    return {
        ("arial", False): normale,
        ("arial", True): grasse,
        # Arial Rounded MT Bold n'existe pas hors Windows : repli sur Arial Bold.
        ("arial rounded mt bold", False): grasse,
        ("arial rounded mt bold", True): grasse,
        ("arial narrow", False): etroite,
        ("arial narrow", True): etroite_grasse,
        ("__defaut__", False): normale,
        ("__defaut__", True): grasse,
    }


def _couleur(couleur_openpyxl, defaut: Color = black) -> Color:
    """Convertit une couleur openpyxl en couleur ReportLab."""
    if couleur_openpyxl is None:
        return defaut
    valeur = getattr(couleur_openpyxl, "rgb", None)
    if not isinstance(valeur, str) or len(valeur) < 6:
        return defaut
    return HexColor("#" + valeur[-6:])


def _largeur_colonne(feuille, index: int) -> float:
    """Largeur d'une colonne, en points."""
    dimension = feuille.column_dimensions.get(get_column_letter(index))
    largeur = dimension.width if dimension and dimension.width else LARGEUR_COLONNE_DEFAUT
    # Conversion Excel : pixels = largeur × 7 + 5 ; points = pixels × 0,75.
    return (largeur * 7 + 5) * 0.75


def _hauteur_ligne(feuille, index: int) -> float:
    dimension = feuille.row_dimensions.get(index)
    return dimension.height if dimension and dimension.height else HAUTEUR_LIGNE_DEFAUT


class Grille:
    """Positions absolues des colonnes et des lignes d'une plage."""

    def __init__(self, feuille, plage: str):
        self.min_col, self.min_row, self.max_col, self.max_row = range_boundaries(plage)
        self.largeurs = {
            col: _largeur_colonne(feuille, col)
            for col in range(self.min_col, self.max_col + 1)
        }
        self.hauteurs = {
            row: _hauteur_ligne(feuille, row)
            for row in range(self.min_row, self.max_row + 1)
        }
        self.largeur = sum(self.largeurs.values())
        self.hauteur = sum(self.hauteurs.values())

    def x(self, colonne: int) -> float:
        """Abscisse du bord gauche de la colonne, relative à la plage."""
        return sum(self.largeurs[c] for c in range(self.min_col, colonne))

    def y(self, ligne: int) -> float:
        """Ordonnée du bord supérieur de la ligne, mesurée vers le bas."""
        return sum(self.hauteurs[r] for r in range(self.min_row, ligne))


def _decouper(texte: str, police: str, taille: float, largeur: float) -> list[str]:
    """Découpe un texte en lignes tenant dans ``largeur``."""
    lignes: list[str] = []
    for paragraphe in texte.split("\n"):
        courante = ""
        for mot in paragraphe.split(" "):
            essai = f"{courante} {mot}".strip()
            if courante and pdfmetrics.stringWidth(essai, police, taille) > largeur:
                lignes.append(courante)
                courante = mot
            else:
                courante = essai
        lignes.append(courante)
    return lignes


class RenduFeuille:
    """Dessine une plage de feuille Excel sur un canevas ReportLab."""

    def __init__(self, feuille, plage: str, valeurs: Optional[dict[str, str]] = None):
        self.feuille = feuille
        self.plage = plage
        self.valeurs = valeurs or {}
        self.grille = Grille(feuille, plage)
        self.polices = enregistrer_polices()
        self._fusions = {}
        for fusion in feuille.merged_cells.ranges:
            gauche, haut, droite, bas = fusion.bounds
            self._fusions[(haut, gauche)] = (bas, droite)
        self._cellules_masquees = {
            (ligne, colonne)
            for fusion in feuille.merged_cells.ranges
            for ligne in range(fusion.bounds[1], fusion.bounds[3] + 1)
            for colonne in range(fusion.bounds[0], fusion.bounds[2] + 1)
        } - set(self._fusions)

    # -- polices ----------------------------------------------------------

    def _police(self, cellule) -> tuple[str, float, float]:
        """Renvoie (police PDF, taille, condensation horizontale en %)."""
        police = cellule.font
        nom = (police.name or "").strip().lower()
        gras = bool(police.bold)
        clef = (nom, gras)
        if clef not in self.polices:
            clef = ("__defaut__", gras)
        nom_pdf = self.polices[clef]
        # Sans véritable fonte étroite, on condense Arial au ratio Arial Narrow.
        condensation = 100.0
        if nom == "arial narrow" and "Narrow" not in nom_pdf:
            condensation = 82.0
        return nom_pdf, float(police.size or 10), condensation

    def _largeur_disponible(self, ligne: int, colonne: int, droite: int,
                            largeur: float) -> float:
        """Largeur utilisable, débordement sur les cellules vides comme dans Excel."""
        suivante = droite + 1
        while suivante <= self.grille.max_col:
            if (ligne, suivante) in self._cellules_masquees \
                    or (ligne, suivante) in self._fusions:
                break
            voisine = self.feuille.cell(row=ligne, column=suivante)
            if self._texte(voisine).strip():
                break
            largeur += self.grille.largeurs[suivante]
            suivante += 1
        return largeur

    # -- texte ------------------------------------------------------------

    def _texte(self, cellule) -> str:
        if cellule.coordinate in self.valeurs:
            return self.valeurs[cellule.coordinate]
        valeur = cellule.value
        if valeur is None or isinstance(valeur, str) and valeur.startswith("="):
            return ""
        return str(valeur)

    def _dessiner_texte(self, canevas, cellule, x, y, largeur, hauteur, echelle,
                        largeur_debordement=None):
        texte = self._texte(cellule)
        if not texte.strip():
            return
        police, taille, condensation = self._police(cellule)
        taille *= echelle
        couleur = _couleur(cellule.font.color)

        marge = 2 * echelle
        alignement = cellule.alignment
        largeur_utile = max(largeur - 2 * marge, 1)

        if alignement.wrap_text:
            lignes = _decouper(texte, police, taille, largeur_utile / (condensation / 100))
        else:
            lignes = [texte]
            # Hors retour à la ligne, Excel déborde sur les cellules vides voisines.
            largeur_utile = max((largeur_debordement or largeur) - 2 * marge, 1)

        # Dernier recours : réduire la police pour qu'aucun texte ne sorte de la page.
        largeur_max = max(
            pdfmetrics.stringWidth(ligne, police, taille) * condensation / 100
            for ligne in lignes
        )
        if largeur_max > largeur_utile:
            taille *= largeur_utile / largeur_max

        interligne = taille * 1.15
        hauteur_texte = interligne * len(lignes)
        vertical = alignement.vertical or "bottom"
        if vertical == "top":
            depart = y + hauteur - taille - marge * 0.5
        elif vertical == "center":
            depart = y + (hauteur + hauteur_texte) / 2 - taille
        else:
            depart = y + marge * 0.5 + interligne * (len(lignes) - 1)

        horizontal = alignement.horizontal or "left"
        for rang, ligne in enumerate(lignes):
            ligne_y = depart - rang * interligne
            largeur_ligne = pdfmetrics.stringWidth(ligne, police, taille) * condensation / 100
            if horizontal == "center":
                ligne_x = x + (largeur - largeur_ligne) / 2
            elif horizontal == "right":
                ligne_x = x + largeur - marge - largeur_ligne
            else:
                ligne_x = x + marge
            objet = canevas.beginText(ligne_x, ligne_y)
            objet.setFont(police, taille)
            objet.setFillColor(couleur)
            if condensation != 100.0:
                objet.setHorizScale(condensation)
            objet.textOut(ligne)
            canevas.drawText(objet)

    # -- bordures et fonds ------------------------------------------------

    def _dessiner_fond(self, canevas, cellule, x, y, largeur, hauteur):
        remplissage = cellule.fill
        if remplissage is None or remplissage.fill_type != "solid":
            return
        couleur = _couleur(remplissage.fgColor, None)
        if couleur is None:
            return
        canevas.setFillColor(couleur)
        canevas.rect(x, y, largeur, hauteur, stroke=0, fill=1)

    def _dessiner_bordures(self, canevas, cellule, x, y, largeur, hauteur, echelle):
        bordure = cellule.border
        canevas.setStrokeColor(black)
        for cote, (x1, y1, x2, y2) in (
            ("left", (x, y, x, y + hauteur)),
            ("right", (x + largeur, y, x + largeur, y + hauteur)),
            ("top", (x, y + hauteur, x + largeur, y + hauteur)),
            ("bottom", (x, y, x + largeur, y)),
        ):
            trait = getattr(bordure, cote)
            if trait is None or not trait.style:
                continue
            epaisseur = 1.2 if trait.style in ("medium", "thick") else 0.6
            canevas.setLineWidth(epaisseur * echelle)
            canevas.line(x1, y1, x2, y2)

    # -- images -----------------------------------------------------------

    def _dessiner_images(self, canevas, origine_x, origine_y, echelle):
        for image in getattr(self.feuille, "_images", []):
            ancre = image.anchor._from
            colonne = ancre.col + 1
            ligne = ancre.row + 1
            if not (self.grille.min_col <= colonne <= self.grille.max_col):
                continue
            largeur = image.width * 0.75 * echelle
            hauteur = image.height * 0.75 * echelle
            x = origine_x + self.grille.x(colonne) * echelle
            y = origine_y - (self.grille.y(ligne) * echelle) - hauteur
            try:
                donnees = image.ref
                lecteur = ImageReader(donnees() if callable(donnees) else donnees)
                canevas.drawImage(lecteur, x, y, largeur, hauteur, mask="auto")
            except Exception:
                # Une image illisible ne doit pas empêcher la production du PDF.
                continue

    # -- rendu ------------------------------------------------------------

    def dessiner(self, canevas, origine_x: float, origine_y: float, echelle: float) -> None:
        """Dessine la plage, ``origine_y`` étant l'ordonnée du bord supérieur."""
        for ligne in range(self.grille.min_row, self.grille.max_row + 1):
            for colonne in range(self.grille.min_col, self.grille.max_col + 1):
                if (ligne, colonne) in self._cellules_masquees:
                    continue
                bas, droite = self._fusions.get((ligne, colonne), (ligne, colonne))
                droite = min(droite, self.grille.max_col)
                bas = min(bas, self.grille.max_row)

                largeur = sum(self.grille.largeurs[c] for c in range(colonne, droite + 1))
                hauteur = sum(self.grille.hauteurs[r] for r in range(ligne, bas + 1))
                x = origine_x + self.grille.x(colonne) * echelle
                y = origine_y - (self.grille.y(ligne) + hauteur) * echelle
                largeur *= echelle
                hauteur *= echelle

                cellule = self.feuille.cell(row=ligne, column=colonne)
                debordement = self._largeur_disponible(
                    ligne, colonne, droite, largeur / echelle) * echelle
                self._dessiner_fond(canevas, cellule, x, y, largeur, hauteur)
                self._dessiner_texte(canevas, cellule, x, y, largeur, hauteur, echelle,
                                     largeur_debordement=debordement)

        # Les bordures sont tracées après les fonds pour rester visibles.
        for ligne in range(self.grille.min_row, self.grille.max_row + 1):
            for colonne in range(self.grille.min_col, self.grille.max_col + 1):
                if (ligne, colonne) in self._cellules_masquees:
                    continue
                bas, droite = self._fusions.get((ligne, colonne), (ligne, colonne))
                droite = min(droite, self.grille.max_col)
                bas = min(bas, self.grille.max_row)
                largeur = sum(self.grille.largeurs[c]
                              for c in range(colonne, droite + 1)) * echelle
                hauteur = sum(self.grille.hauteurs[r]
                              for r in range(ligne, bas + 1)) * echelle
                x = origine_x + self.grille.x(colonne) * echelle
                y = origine_y - (self.grille.y(ligne) * echelle) - hauteur
                cellule = self.feuille.cell(row=ligne, column=colonne)
                self._dessiner_bordures(canevas, cellule, x, y, largeur, hauteur, echelle)

        self._dessiner_images(canevas, origine_x, origine_y, echelle)


def rendre_page(feuille, plage: str, destination, valeurs: Optional[dict[str, str]] = None,
                titre: str = "") -> Path:
    """Produit un PDF A4 portrait d'une seule page contenant ``plage``."""
    rendu = RenduFeuille(feuille, plage, valeurs)
    largeur_page, hauteur_page = A4

    marges = feuille.page_margins
    marge_gauche = float(marges.left or 0) * 72
    marge_droite = float(marges.right or 0) * 72
    marge_haut = float(marges.top or 0) * 72
    marge_bas = float(marges.bottom or 0) * 72

    utile_largeur = largeur_page - marge_gauche - marge_droite
    utile_hauteur = hauteur_page - marge_haut - marge_bas
    echelle = min(utile_largeur / rendu.grille.largeur,
                  utile_hauteur / rendu.grille.hauteur)

    chemin = Path(destination)
    chemin.parent.mkdir(parents=True, exist_ok=True)
    canevas = rl_canvas.Canvas(str(chemin), pagesize=A4)
    canevas.setTitle(titre or chemin.stem)
    rendu.dessiner(canevas, marge_gauche, hauteur_page - marge_haut, echelle)
    canevas.showPage()
    canevas.save()
    return chemin
